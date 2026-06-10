import os
import json
import re
import joblib
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer

TEMPLATE_URLS = {
    'about.html': ('/about', 'Our Mission'),
    'academy.html': ('/academy', 'Academy Quiz'),
    'activities.html': ('/activities', 'Volunteer / Activities'),
    'calculator.html': ('/calculator', 'Impact Calculator'),
    'contact.html': ('/contact', 'Contact Us'),
    'corporate.html': ('/corporate-partnerships', 'Corporate ESG Portal'),
    'country_detail.html': ('/country/kenya', 'Country Details'),
    'faq.html': ('/faq', 'FAQ'),
    'forest_audit.html': ('/forest-audit', 'Forest Audit Map'),
    'impact.html': ('/impact', 'Transparency & Impact'),
    'index.html': ('/', 'Home'),
    'press_news.html': ('/press-news', 'Press Room'),
    'privacy.html': ('/privacy', 'Privacy Policy'),
    'resources.html': ('/resources', 'Resources'),
    'services.html': ('/services', 'Our Services'),
    'team.html': ('/our-team', 'The Team'),
    'terms.html': ('/terms', 'Terms of Service'),
    'work.html': ('/work', 'Careers / Work with Us'),
    'blog.html': ('/blog', 'Impact Stories'),
    'events.html': ('/events', 'Events'),
    'pay.html': ('/gift-a-tree', 'Gift a Tree'),
    'success.html': ('/donation-success', 'Donation Success'),
    'register.html': ('/register', 'Register'),
}

def clean_text(text):
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def clean_html_content(html_text):
    # Remove script and style blocks
    html_text = re.sub(r'<script.*?>.*?</script>', ' ', html_text, flags=re.DOTALL | re.IGNORECASE)
    html_text = re.sub(r'<style.*?>.*?</style>', ' ', html_text, flags=re.DOTALL | re.IGNORECASE)
    # Remove Jinja tags and comments
    html_text = re.sub(r'\{#.*?#\}', ' ', html_text, flags=re.DOTALL)
    html_text = re.sub(r'\{%.*?%\}', ' ', html_text, flags=re.DOTALL)
    html_text = re.sub(r'\{\{.*?\}\}', ' ', html_text, flags=re.DOTALL)
    # Remove HTML comments and tags
    html_text = re.sub(r'<!--.*?-->', ' ', html_text, flags=re.DOTALL)
    html_text = re.sub(r'<[^>]+>', ' ', html_text)
    # Normalize whitespace
    html_text = re.sub(r'\s+', ' ', html_text)
    return html_text.strip()

def chunk_text(text, max_len=400, min_len=50):
    sentences = re.split(r'(?<=[.!?])\s+', text)
    chunks = []
    curr = ""
    for s in sentences:
        s = s.strip()
        if not s:
            continue
        if len(curr) + len(s) + 1 <= max_len:
            if curr:
                curr += " " + s
            else:
                curr = s
        else:
            if len(curr) >= min_len:
                chunks.append(curr)
            curr = s
    if len(curr) >= min_len:
        chunks.append(curr)
    return chunks

def extract_pdf_text(file_path):
    from pypdf import PdfReader
    print(f"Reading PDF: {os.path.basename(file_path)}")
    reader = PdfReader(file_path)
    text = ""
    total_pages = len(reader.pages)
    for idx, page in enumerate(reader.pages):
        if idx % 50 == 0:
            print(f"  page {idx}/{total_pages}")
        try:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        except Exception as e:
            print(f"  Error page {idx}: {e}")
    return text

def extract_docx_text(file_path):
    from docx import Document
    print(f"Reading DOCX: {os.path.basename(file_path)}")
    doc = Document(file_path)
    text = ""
    for para in doc.paragraphs:
        if para.text.strip():
            text += para.text + "\n"
    return text

def extract_xlsx_text(file_path):
    from openpyxl import load_workbook
    print(f"Reading XLSX: {os.path.basename(file_path)}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_str = " ".join([str(cell) for cell in row if cell is not None])
            if row_str.strip():
                text += row_str + "\n"
    return text

def extract_country_data(main_py_path):
    import ast
    if not os.path.exists(main_py_path):
        print(f"Error: {main_py_path} not found.")
        return {}
    try:
        with open(main_py_path, 'r', encoding='utf-8') as f:
            node = ast.parse(f.read())
        for element in node.body:
            if isinstance(element, ast.Assign):
                for target in element.targets:
                    if isinstance(target, ast.Name) and target.id == 'COUNTRY_DATA':
                        return ast.literal_eval(element.value)
    except Exception as e:
        print(f"Error parsing main.py statically for COUNTRY_DATA: {e}")
    return {}

def build_knowledge_base(docs_dir, templates_dir, main_py_path, output_json):
    print(f"Scanning documents in {docs_dir} and templates in {templates_dir}...")
    knowledge_base = {
        "company_info": {
            "name": "TEAMEnvironment KENYA",
            "motto": "Build a Heritage for Future Generations of a Cleaner, Greener (Food & Water) Secure and a Peaceful Environment.",
            "website": "teamenvironment.org",
            "contacts": ["+254 718 052745"],
            "email": "teamenvironment.ke@gmail.com",
            "location": "Nairobi, Kenya"
        },
        "documents": []
    }

    # 1. Parse HTML templates
    if os.path.exists(templates_dir):
        for root_dir, dirs, files in os.walk(templates_dir):
            if "admin" in root_dir.split(os.sep):
                continue
            for file in files:
                if not file.endswith('.html'):
                    continue
                if file in ['receipt_template.html', 'pesapal_simulator.html', 'base.html']:
                    continue
                
                file_path = os.path.join(root_dir, file)
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        html_content = f.read()
                    
                    content_match = re.search(r'\{%\s*block\s+content\s*%\}(.*?)\{%\s*endblock\s*%\}', html_content, flags=re.DOTALL)
                    if content_match:
                        raw_html = content_match.group(1)
                    else:
                        raw_html = html_content
                    
                    clean_text_str = clean_html_content(raw_html)
                    if not clean_text_str:
                        continue
                    
                    chunks = chunk_text(clean_text_str)
                    if chunks:
                        url, title = TEMPLATE_URLS.get(file, (f"/{file}", file.replace('.html', '').capitalize()))
                        knowledge_base["documents"].append({
                            "title": title,
                            "url": url,
                            "content": chunks
                        })
                        print(f"  Processed template: {file} -> {url} ({len(chunks)} chunks)")
                except Exception as e:
                    print(f"  Error reading template {file_path}: {e}")

    # 2. Parse COUNTRY_DATA from main.py
    country_data = extract_country_data(main_py_path)
    if country_data:
        for key, info in country_data.items():
            name = info.get('name')
            region = info.get('region')
            trees = info.get('trees')
            volunteers = info.get('volunteers')
            acres = info.get('acres')
            projects = info.get('projects')
            desc = info.get('description')
            goals = info.get('goals')
            
            paragraph = f"In {name} ({region}), TEAMEnvironment is active with {trees} trees planted, {volunteers} volunteers, {acres} acres restored, and {projects} projects. {desc} The goals for {name} are: {goals}"
            
            knowledge_base["documents"].append({
                "title": region,
                "url": f"/country/{key}",
                "content": [paragraph]
            })
            print(f"  Processed Country: {name} -> /country/{key}")

    # 3. Parse PDFs, DOCX, XLSX
    if os.path.exists(docs_dir):
        for filename in os.listdir(docs_dir):
            file_path = os.path.join(docs_dir, filename)
            if os.path.isdir(file_path):
                continue

            ext = os.path.splitext(filename)[1].lower()
            extracted_text = ""
            
            try:
                if ext == ".pdf":
                    extracted_text = extract_pdf_text(file_path)
                elif ext == ".docx":
                    extracted_text = extract_docx_text(file_path)
                elif ext == ".xlsx":
                    extracted_text = extract_xlsx_text(file_path)
                else:
                    print(f"  Skipping unsupported file type: {filename}")
                    continue
            except Exception as e:
                print(f"  Failed to extract text from {filename}: {e}")
                continue

            if not extracted_text.strip():
                print(f"  No text extracted from {filename}.")
                continue

            # Clean and chunk
            clean_text_str = clean_text(extracted_text)
            valid_chunks = chunk_text(clean_text_str)
            
            print(f"  Extracted {len(valid_chunks)} chunks from {filename}")
            
            knowledge_base["documents"].append({
                "title": filename,
                "url": f"/static/documents/{filename}",
                "content": valid_chunks
            })

    os.makedirs(os.path.dirname(output_json), exist_ok=True)
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(knowledge_base, f, indent=4, ensure_ascii=False)
    
    print(f"Saved knowledge base to {output_json}")
    return knowledge_base

def train_model(json_path, model_dir):
    print("Training TF-IDF chatbot model...")
    if not os.path.exists(json_path):
        print(f"Error: {json_path} not found.")
        return

    with open(json_path, 'r', encoding='utf-8') as f:
        kb = json.load(f)

    data = []
    for doc in kb.get("documents", []):
        for paragraph in doc.get("content", []):
            data.append({
                "source": doc.get("title", ""),
                "url": doc.get("url", ""),
                "text": paragraph
            })

    if not data:
        print("Warning: No document text content to train on.")
        return

    df = pd.DataFrame(data)
    
    # Train Vectorizer
    vectorizer = TfidfVectorizer(stop_words='english')
    tfidf_matrix = vectorizer.fit_transform(df['text'].astype(str))

    # Save artifacts
    os.makedirs(model_dir, exist_ok=True)
    joblib.dump(vectorizer, os.path.join(model_dir, "tfidf_vectorizer.pkl"))
    joblib.dump(tfidf_matrix, os.path.join(model_dir, "document_vectors.pkl"))
    joblib.dump(df, os.path.join(model_dir, "knowledge_df.pkl"))
    
    print(f"Training completed successfully. Artifacts saved to {model_dir}")

if __name__ == "__main__":
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    DOCS_DIR = os.path.join(os.path.dirname(BASE_DIR), "static", "documents")
    TEMPLATES_DIR = os.path.join(os.path.dirname(BASE_DIR), "templates")
    MAIN_PY_PATH = os.path.join(os.path.dirname(BASE_DIR), "main.py")
    JSON_PATH = os.path.join(BASE_DIR, "data", "knowledge.json")
    MODEL_DIR = os.path.join(BASE_DIR, "model")
    
    build_knowledge_base(DOCS_DIR, TEMPLATES_DIR, MAIN_PY_PATH, JSON_PATH)
    train_model(JSON_PATH, MODEL_DIR)
